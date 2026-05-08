import pandas as pd
import glob
from fpdf import FPDF
from pathlib import Path
import openpyxl

filepath = glob.glob("Invoice/*.xlsx")

for file_ded in filepath:
    df = pd.read_excel(file_ded, sheet_name= "Sheet 1")
    pdf = FPDF(orientation="P", unit= "mm", format= "A4")
    pdf.add_page()
    filename = Path(file_ded).stem
    invoice_nr = filename.split("-")[0]
    pdf.set_font(family="Times", size= 16, style= "B")
    pdf.cell(w=50, h=8, txt= f"Invoice name:{invoice_nr}", ln=1)

    time_Date = filename.split("-")[1]
    pdf.set_font(family="Times", size= 16, style= "B")
    pdf.cell(w=50, h=8, txt= f"Date:{time_Date}", ln=1)

    col = list(df.columns)
    pdf.set_font(family="Times", size = 10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=col[0], border=1)
    pdf.cell(w=70, h=8, txt=col[1], border=1)
    pdf.cell(w=70, h=8, txt=col[2], border=1)
    pdf.cell(w=70, h=8, txt=col[3], border=1)
    pdf.cell(w=70, h=8, txt=col[4], border=1, ln=1)

    for index, row in df.iterrows():
        pdf.set_font(family="Times", size = 10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["total_price"]), border=1, ln=1)


    pdf.output(f"Pdf/{invoice_nr}.pdf")